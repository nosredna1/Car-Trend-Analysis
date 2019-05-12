import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.cbook as cbook
import numpy as np 
from openpyxl import load_workbook, Workbook

def sort(dateList, pricingList):
    for i in range(1, len(dateList)):
        j = i - 1
        nextE = dateList[i]
        nextE_pricing = pricingList[i]

        while (dateList[j] > nextE) and (j >= 0):
            dateList[j + 1] = dateList[j]
            pricingList[j + 1] = pricingList[j]
            j = j - 1
        dateList[j + 1] = nextE
        pricingList[j + 1] = nextE_pricing
    
    return (dateList, pricingList)

def extractData(worksheet):
    date = []
    pricing = []

    nRows = worksheet.max_row
    # Get titles of listing from workbook input
    for r in range(nRows - 1):
        date_time = worksheet.cell(row = r + 2, column = 1).value
        date_day = date_time.split()
        date.append(date_day[0])
        pricing.append(worksheet.cell(row = r + 2, column = 2).value)

    (date, pricing) = sort(date, pricing)

    return (date, pricing)

def pricing_plot(date, pricing):

    fig, ax = plt.subplots()
    ax.scatter(date, pricing)
    
    fig.autofmt_xdate()

    plt.xlabel('Listing Date')
    plt.ylabel('Listed Price ($)')
    ax.grid(True)
    plt.show()

if __name__ == "__main__":
    loaded_workbook = load_workbook(filename = 'ScrappedListings.xlsx')
    loaded_worksheet = loaded_workbook.active

    (date, pricing) = extractData(loaded_worksheet)
    pricing_plot(date, pricing)