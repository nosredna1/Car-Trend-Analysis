from requests import get
from requests.exceptions import RequestException
from contextlib import closing
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import os.path, xlsxwriter

def simple_get(url):
# Get content by user input url by making an HTTP GET request.
# Content type can be HTML/XML, check to see which one

    try:
        with closing(get(url, stream=True)) as resp:
            if is_good_response(resp):
                return resp.content
            else:
                return None

    except RequestException as e:
        log_error('Error during re quests to {0} : {1}'.formart(url, str(e)))
        return None

def is_good_response(resp):
# Determines whether response if HTML

    content_type = resp.headers['Content-Type'].lower()
    return (resp.status_code == 200
            and content_type is not None
            and content_type.find('html') > -1)

def log_error(e):
    print(e)

def get_info(listings):
    car_info_listings = []

    for i in range(len(listings)):
        title = listings[i].find('a', class_= 'result-title hdrlnk').text
        pricing = int(listings[i].find('span', class_= 'result-price').text.replace('$',''))
        date = listings[i].find('time', class_= 'result-date')['datetime']
        id = int(listings[i].find('a', class_= 'result-title hdrlnk')['data-id'])
        link = listings[i].find('a', class_= 'result-title hdrlnk')['href']
        
        car_info = [title, pricing, date, id, link]
        car_info_listings.append(car_info)

    return(car_info_listings)

def get_listings(raw_input):
    major_cities = ['atlanta', 'houston', 'losangeles', 'sfbay', 'chicago', 'newyork',
            'seattle', 'orangecounty', 'sandiego', 'washingtondc', 'portland', 'boston',
            'phoenix', 'denver']

    # Download the searched page of listed cars
    url_pre = 'https://'
    url_end = '.craigslist.org/search/cta?query='
    has_image = '&sort=rel&hasPic=1'

    # Checks for spaces and replaces with +
    if " " in raw_input:
        searched_car = raw_input.replace(" ", "+")
    else:
        searched_car = raw_input

    parsed_listings = []
    for city in major_cities:
        # Create standard url for searching specific vehicle
        combined_url = url_pre + city + url_end + searched_car + has_image

        # Obtains HTML response and checks if a response works
        response = simple_get(combined_url)
        if response is not None:
            html = BeautifulSoup(response, 'html.parser')
            listings = (html.find_all('li', class_= 'result-row'))
            parsed_infos = get_info(listings)

            # Extract parsed data into a single list
            # car_info = [title, pricing, date, id, link]
            for single_listing in parsed_infos:
                parsed_listings.append(single_listing)
        
    return parsed_listings

def filtered_search(raw_input, car_info):
    # Splits keywords into individual words
    car_model = raw_input.split()
    current_len = len(car_info)
    i = 0
    for k in car_model:
        while i < current_len:
            if k.upper() in car_info[i][0].upper():
                i += 1
            else:
                # Removes listings that does not contain searched keywords
                del car_info[i]

                current_len = len(car_info)
        i = 0 # Reset counter

    return(car_info)

def createNewWorksheet(car_info):
    # Creating a workbook and worksheet if no previous existed
    workbook = xlsxwriter.Workbook('ScrappedListings.xlsx')
    worksheet = workbook.add_worksheet('Listings')

    # Adding bold, money formats
    bold = workbook.add_format({'bold': True})
    money = workbook.add_format({'num_format': '$#,##0'})

    # Data Headers
    worksheet.write('A1', 'Date', bold)
    worksheet.write('B1', 'Listed Price', bold)
    worksheet.write('C1', 'ID', bold)
    worksheet.write('D1', 'Title', bold)
    worksheet.write('E1', 'Link', bold)

    # Starting from first cell below headers
    row = 1
    col = 0

    # Writes all scrapped info into specific columns
    for i in range(len(car_info)):
        worksheet.write(row, col, car_info[i][2])
        worksheet.write(row, col + 1, car_info[i][1], money)
        worksheet.write(row, col + 2, car_info[i][3])
        worksheet.write(row, col + 3, car_info[i][0])
        worksheet.write(row, col + 4, car_info[i][4])
        row += 1

    workbook.close()

def addListings(working_worksheet, car_info):
    # Determines how many rows are written in the excel sheet and
    # adds the amount of new listings to the list
    nRows = working_worksheet.max_row
    nAdd = len(car_info[0])

    i = 0
    for r in range(nRows, nRows + nAdd):
        working_worksheet.cell(row = r, column = 1).value = car_info[i][2]
        working_worksheet.cell(row = r, column = 2).value  = car_info[i][1]
        working_worksheet.cell(row = r, column = 3).value  = car_info[i][3]
        working_worksheet.cell(row = r, column = 4).value  = car_info[i][0]
        working_worksheet.cell(row = r, column = 5).hyperlink  = car_info[i][4]

        working_worksheet.cell(row = r, column = 2).number_format = u'"$ "#,##'
        working_worksheet.cell(row = r, column = 5).style  = 'Hyperlink'

        i += 1

    loaded_workbook.save('ScrappedListings.xlsx')
    loaded_workbook.close()

def checkDup(car_info, workbook = None):
    # Removes redudant listings made by dealerships usually
    if workbook is not None:
        workbook_car_info = []
        nRows = workbook.max_row
        # Get titles of listing from workbook input
        for r in range(nRows - 1):
            workbook_car_info.append(workbook.cell(row = r + 2, column = 4).value)

        # Compare workbook titles with input newly scraped car infos
        x = 0
        y = 0
        while x < nRows - 1: # First loop iterates workbook infos
            # Have to update list length as you remove
            current_car_listings = len(car_info)
            
            while y < current_car_listings: # Second loop iterates over scraped car infos
                if workbook_car_info[x] is car_info[y][0]:
                    del car_info[y]

                    current_car_listings = len(car_info)
                else:
                    y += 1
            x += 1
            y = 0
        # At this point, all listings already in the workbook should remove duplicates
        # found in newly scraped listings        

    # This portion removes duplicates found within the newly scrapped listings
    z = 0
    k = 0
    new_car_listings = len(car_info)
    while z < new_car_listings:
        selected_listing = car_info[z][0]

        while k < new_car_listings:
            if k != z and selected_listing is car_info[k][0]:
                del car_info[k]
            else:
                k += 1
        z += 1
        k = 0

    return car_info

if __name__ == "__main__":
    # Takes user input of 
    raw_input = input('What vehicle? ')

    # Rips listings and puts them into a list consisting of [title, pricing, date, id, link]
    car_info = get_listings(raw_input)
    upCar_info = filtered_search(raw_input, car_info)

    # Checks if a file exists with old ripped listings
    exists = os.path.isfile('ScrappedListings.xlsx')

    if exists:
        loaded_workbook = load_workbook(filename = 'ScrappedListings.xlsx')
        loaded_worksheet = loaded_workbook.active
        filtered_car_info = checkDup(upCar_info, loaded_worksheet)
        addListings(loaded_worksheet, filtered_car_info)
    else:
        filtered_car_info = checkDup(upCar_info)
        createNewWorksheet(filtered_car_info)
